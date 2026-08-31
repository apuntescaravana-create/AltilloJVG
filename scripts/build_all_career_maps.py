"""
build_all_career_maps.py
Genera las 18 plantillas oficiales de "Mapa de Carrera Personal" (.xlsx)
para todos los profesorados del ISP Joaquín V. González.
Elimina completamente comisiones duplicadas ("A", "B", "1°C", etc.)
dejando únicamente las materias oficiales únicas de cada plan de estudios.
"""
import os
import json
import re
import unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = r"d:\IA\AltilloJVG"
OUTPUT_DIR = os.path.join(BASE_DIR, "data", "mapas")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 18 Carreras Oficiales
CAREER_FILENAMES = {
    "Profesorado de Filosofía": "Mapa_de_Carrera_Filosofia.xlsx",
    "Profesorado de Historia": "Mapa_de_Carrera_Historia.xlsx",
    "Profesorado de Lengua y Literatura": "Mapa_de_Carrera_Lengua_y_Literatura.xlsx",
    "Profesorado de Inglés": "Mapa_de_Carrera_Ingles.xlsx",
    "Profesorado de Matemática": "Mapa_de_Carrera_Matematica.xlsx",
    "Profesorado de Biología": "Mapa_de_Carrera_Biologia.xlsx",
    "Profesorado de Física": "Mapa_de_Carrera_Fisica.xlsx",
    "Profesorado de Química": "Mapa_de_Carrera_Quimica.xlsx",
    "Profesorado de Geografía": "Mapa_de_Carrera_Geografia.xlsx",
    "Profesorado de Informática": "Mapa_de_Carrera_Informatica.xlsx",
    "Profesorado de Psicología": "Mapa_de_Carrera_Psicologia.xlsx",
    "Profesorado de Ciencias de la Educación": "Mapa_de_Carrera_Ciencias_de_la_Educacion.xlsx",
    "Profesorado de Cs. de la Educación": "Mapa_de_Carrera_Ciencias_de_la_Educacion.xlsx",
    "Profesorado de Ciencias Jurídicas": "Mapa_de_Carrera_Ciencias_Juridicas.xlsx",
    "Profesorado de Cs. Jurídicas": "Mapa_de_Carrera_Ciencias_Juridicas.xlsx",
    "Profesorado de Ciencia Política": "Mapa_de_Carrera_Ciencia_Politica.xlsx",
    "Profesorado de Economía": "Mapa_de_Carrera_Economia.xlsx",
    "Profesorado de Ciencias de la Administración": "Mapa_de_Carrera_Ciencias_de_la_Administracion.xlsx",
    "Profesorado de Francés": "Mapa_de_Carrera_Frances.xlsx",
    "Profesorado de Italiano": "Mapa_de_Carrera_Italiano.xlsx"
}

def normalize_key(s):
    nfkd = unicodedata.normalize('NFKD', s.lower())
    clean = ''.join([c for c in nfkd if not unicodedata.combining(c)])
    clean = re.sub(r'[^a-z0-9]', '', clean)
    return clean

def clean_subject_name(s):
    # 1. Normalizar LEO
    if re.search(r'lectura\s*,\s*escritura\s*y\s*oralidad\s*2|\bLEO\s*2\b', s, re.I):
        return "Lectura, Escritura y Oralidad II (LEO 2)"
    if re.search(r'lectura\s*,\s*escritura\s*y\s*oralidad|\bLEO\s*1\b|\bleo\b|lectura\s+y\s+escritura\s+acad', s, re.I):
        return "Lectura, Escritura y Oralidad I (LEO 1)"

    # 2. Remover comisiones entre comillas ("A", "B", "1°C", "T", "H", etc.)
    s = re.sub(r'[\"\'\“\”][^\"]*?[\"\'\“\”]', '', s)

    # 3. Remover anotaciones de cuatrimestre y paréntesis residuales
    s = re.sub(r'\(\s*\d+°?\s*(?:cuat|cuatrimestre|semestre|c)\.?\s*\)', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\(\s*(?:cuat|cuatrimestre|semestre|c)\.?\s*\)', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\b\d+°?\s*(?:cuat|cuatrimestre|c)\b\.?', '', s, flags=re.IGNORECASE)

    # 4. Remover sufijos de comisiones con guiones o barras (- A, / B)
    s = re.sub(r'[\-\/\–]\s*[A-Za-z0-9]\b', '', s)

    # 5. Remover letras sueltas de comisiones al final (ej: PEDAGOGIA A, QUIMICA B) EXCEPTUANDO números romanos
    s = re.sub(r'\s+[A-HJ-UW-Z]\b$', '', s)
    s = re.sub(r'\s+[a-hj-uw-z]\b$', '', s)

    # 6. Normalizar abreviaturas a sus nombres oficiales completos
    s = re.sub(r'\bHist\.\b|\bHist\b', 'Historia', s, flags=re.IGNORECASE)
    s = re.sub(r'\bIntrod\.\b|\bIntro\.\b|\bIntrod\b|\bIntro\b', 'Introducción', s, flags=re.IGNORECASE)
    s = re.sub(r'\bGeog\.\b|\bGeog\b', 'Geografía', s, flags=re.IGNORECASE)
    s = re.sub(r'\bCult\.\b|\bCult\b', 'Cultura', s, flags=re.IGNORECASE)
    s = re.sub(r'\bBrit\.\b|\bBritanicas\b|\bBritanica\b', 'Británicas', s, flags=re.IGNORECASE)
    s = re.sub(r'\bEE\.UU\.\b|\bEe\.uu\b|\bEEUU\b|\bEe\.Uu\b', 'EE.UU.', s, flags=re.IGNORECASE)
    s = re.sub(r'\bCs\.\b|\bCs\b', 'Ciencias', s, flags=re.IGNORECASE)
    s = re.sub(r'\bDidact\.\b|\bDidact\b', 'Didáctica', s, flags=re.IGNORECASE)
    s = re.sub(r'\bLab\.\b|\bLab\b', 'Laboratorio', s, flags=re.IGNORECASE)
    s = re.sub(r'\bDd\s*Hh\b|\bDdhh\b|\bDd\.Hh\.\b', 'Derechos Humanos', s, flags=re.IGNORECASE)
    s = re.sub(r'\bConst\.?\s*(?:de\s+la\s+)?pract\.?\s*doc\.?\b|\bConstr\.?\s*practica\s*docente\b|\bConstruccion\s+de\s+la\s+pract\.?\s*docente\b', 'Construcción de la Práctica Docente', s, flags=re.IGNORECASE)
    s = re.sub(r'\bNiv\.?\s*inic\.?\s*y\s*prim\.?\b|\bN\s*Inic\s*y\s*Prim\b|\bNivel\s*Inicial\s*y\s*Primario\b|\bNiv\.?\s*Inic\.?\s*y\s*Pri\b', 'en Nivel Inicial y Primario', s, flags=re.IGNORECASE)
    s = re.sub(r'\bNiv\.?\s*medio\b|\bN\.?\s*Medio\b|\bNivel\s*Medio\b', 'en Nivel Medio', s, flags=re.IGNORECASE)
    s = re.sub(r'\by\s*Res\.?\b|\by\s*Residencia\b', 'y Residencia', s, flags=re.IGNORECASE)
    s = re.sub(r'\bMorfofisiolog\s+comp\b', 'Morfofisiología Comparada', s, flags=re.IGNORECASE)
    s = re.sub(r'\by\s*Amb\b', 'y Ambiente', s, flags=re.IGNORECASE)
    s = re.sub(r'\bSist\.\b', 'Sistemas', s, flags=re.IGNORECASE)
    s = re.sub(r'\bDesarrollo\s+y\s+Aprend\.\b', 'Desarrollo y Aprendizaje', s, flags=re.IGNORECASE)

    # 7. Acentuación estándar y ortografía
    replacements = [
        (r'\bFonetica\b', 'Fonética'), (r'\bFonologia\b', 'Fonología'),
        (r'\bGramatica\b', 'Gramática'), (r'\bIngles\b', 'Inglés'),
        (r'\bPedagogia\b', 'Pedagogía'), (r'\bDidactica\b', 'Didáctica'),
        (r'\bMatematica\b', 'Matemática'), (r'\bAnalisis\b', 'Análisis'),
        (r'\bAlgebra\b', 'Álgebra'), (r'\bFisica\b', 'Física'),
        (r'\bQuimica\b', 'Química'), (r'\bFilosofia\b', 'Filosofía'),
        (r'\bPsicologia\b', 'Psicología'), (r'\bSociologia\b', 'Sociología'),
        (r'\bEducacion\b', 'Educación'), (r'\bEspana\b', 'España'),
        (r'\bEspanol\b', 'Español'), (r'\bEtica\b', 'Ética'),
        (r'\bAdquisicion\b', 'Adquisición'), (r'\bLingüistica\b|\bLinguistica\b', 'Lingüística'),
        (r'\bEducativa\b', 'Educativa'), (r'\bInformatica\b', 'Informática'),
        (r'\bBiologia\b', 'Biología'), (r'\bGeografia\b', 'Geografía'),
        (r'\bEconomia\b', 'Economía'), (r'\bAdministracion\b', 'Administración'),
        (r'\bJuridicas\b', 'Jurídicas'), (r'\bPolitica\b', 'Política'),
        (r'\bFrances\b', 'Francés'), (r'\bBasica\b|\bBasico\b|\bBasicos\b', 'Básica'),
        (r'\bTeoria\b', 'Teoría'), (r'\bPractica\b', 'Práctica'),
        (r'\bAcreditacion\b', 'Acreditación'), (r'\bEspecializacion\b', 'Especialización'),
        (r'\bInvestigacion\b', 'Investigación')
    ]
    for pattern, repl in replacements:
        s = re.sub(pattern, repl, s, flags=re.IGNORECASE)

    # 8. Limpiar espacios redundantes y puntuación residual
    s = re.sub(r'\s+', ' ', s).strip(' -–/,.:;')

    # 9. Capitalización elegante (Title Case preservando conectores y números romanos)
    if s.isupper() or s.islower():
        words = s.split()
        title_words = []
        for w in words:
            if re.match(r'^(I|II|III|IV|V|VI|VII|VIII|IX|X|XI|XII)$', w, re.IGNORECASE):
                title_words.append(w.upper())
            elif w.lower() in ['de', 'del', 'la', 'las', 'el', 'los', 'y', 'en', 'a', 'para', 'por', 'con', 'e', 'o', 'al']:
                title_words.append(w.lower())
            elif w.upper() in ['EE.UU.', 'AYTP', 'LEO', 'EDI', 'TIC', 'TICS', 'ESI']:
                title_words.append(w.upper())
            else:
                title_words.append(w.capitalize())
        if title_words:
            title_words[0] = title_words[0].capitalize()
        s = ' '.join(title_words)

    return s

def build_career_excel(career_name, raw_years_data, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapa de Carrera"

    # Paleta de Colores Institucional
    navy_fill = PatternFill(start_color="0B2545", end_color="0B2545", fill_type="solid")
    cyan_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    gold_fill = PatternFill(start_color="D99B26", end_color="D99B26", fill_type="solid")
    light_blue_fill = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
    light_gray_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    section_fill = PatternFill(start_color="134074", end_color="134074", fill_type="solid")
    green_soft = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

    font_title = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Arial", size=10, italic=True, color="BAE6FD")
    font_section = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_regular = Font(name="Arial", size=10)
    font_bold = Font(name="Arial", size=10, bold=True)
    font_indicator = Font(name="Arial", size=11, bold=True, color="0B2545")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # 1. Encabezado Institucional
    ws.merge_cells("A1:G1")
    ws["A1"] = f"MAPA DE CARRERA | {career_name.upper()}"
    ws["A1"].font = font_title
    ws["A1"].fill = navy_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:G2")
    ws["A2"] = "La Caravana + Estudiantes Independientes | Altillo JVG - ISP Joaquín V. González"
    ws["A2"].font = font_subtitle
    ws["A2"].fill = navy_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # 2. Encabezados de Columnas de Materias
    headers = [
        "Materia / Asignatura Oficial",
        "Cursada Regular",
        "Final Aprobado",
        "Año / Tramo",
        "Condición Cursada",
        "Condición Final",
        "Estado General"
    ]
    ws.row_dimensions[4].height = 26
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = cyan_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Validación de Datos (SI, NO)
    dv = DataValidation(type="list", formula1='"SI,NO"', allow_blank=True)
    dv.error = "Por favor, elegí 'SI' o 'NO' de la lista desplegable."
    dv.errorTitle = "Opción inválida"
    ws.add_data_validation(dv)

    current_row = 5
    materia_rows = []
    seen_in_career = set()

    for year_name, subjects in raw_years_data.items():
        clean_subjects_for_year = []
        for s in subjects:
            clean_name = clean_subject_name(s)
            if not clean_name:
                continue
            key = normalize_key(clean_name)
            if key in seen_in_career:
                continue
            seen_in_career.add(key)
            clean_subjects_for_year.append(clean_name)

        if not clean_subjects_for_year:
            continue

        # Fila de Sección (Año)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        section_cell = ws.cell(row=current_row, column=1, value=f"📌 {year_name.upper()}")
        section_cell.font = font_section
        section_cell.fill = section_fill
        section_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for subj in clean_subjects_for_year:
            row = current_row
            materia_rows.append(row)

            # Col A: Nombre de Materia
            c_a = ws.cell(row=row, column=1, value=subj)
            c_a.font = font_regular
            c_a.border = thin_border
            c_a.alignment = Alignment(horizontal="left", vertical="center", indent=1)

            # Col B: Cursada Regular (Default NO)
            c_b = ws.cell(row=row, column=2, value="NO")
            c_b.font = font_bold
            c_b.alignment = Alignment(horizontal="center", vertical="center")
            c_b.border = thin_border
            c_b.fill = light_blue_fill
            dv.add(c_b)

            # Col C: Final Aprobado (Default NO)
            c_c = ws.cell(row=row, column=3, value="NO")
            c_c.font = font_bold
            c_c.alignment = Alignment(horizontal="center", vertical="center")
            c_c.border = thin_border
            c_c.fill = light_blue_fill
            dv.add(c_c)

            # Col D: Año
            c_d = ws.cell(row=row, column=4, value=year_name)
            c_d.font = font_regular
            c_d.alignment = Alignment(horizontal="center", vertical="center")
            c_d.border = thin_border

            # Col E: Condición Cursada (Fórmula)
            c_e = ws.cell(row=row, column=5, value=f'=IF(B{row}="SI", "REGULAR", "PENDIENTE")')
            c_e.font = font_regular
            c_e.alignment = Alignment(horizontal="center", vertical="center")
            c_e.border = thin_border

            # Col F: Condición Final (Fórmula)
            c_f = ws.cell(row=row, column=6, value=f'=IF(C{row}="SI", "APROBADO", IF(B{row}="SI", "ADEUDA FINAL", "ADEUDA"))')
            c_f.font = font_regular
            c_f.alignment = Alignment(horizontal="center", vertical="center")
            c_f.border = thin_border

            # Col G: Estado General (Fórmula con Emoji)
            c_g = ws.cell(row=row, column=7, value=f'=IF(C{row}="SI", "🟢 Aprobada", IF(B{row}="SI", "🟡 Regular", "🟣 Pendiente"))')
            c_g.font = font_bold
            c_g.alignment = Alignment(horizontal="center", vertical="center")
            c_g.border = thin_border

            ws.row_dimensions[row].height = 20
            current_row += 1

    # 3. Panel Lateral de Estadísticas y Control de Trayectoria
    total_materias = len(materia_rows)
    if total_materias > 0:
        first_row = materia_rows[0]
        last_row = materia_rows[-1]

        ws["I1"] = "CONTROL DE TRAYECTORIA"
        ws["I1"].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        ws["I1"].fill = navy_fill
        ws["I1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("I1:K1")

        stat_items = [
            ("Total de Materias del Plan", total_materias, False),
            ("Materias con Cursada Regular", f'=COUNTIF(B{first_row}:B{last_row}, "SI")', False),
            ("Finales / Materias Aprobadas", f'=COUNTIF(C{first_row}:C{last_row}, "SI")', False),
            ("Finales Adeudados", f'=COUNTIF(F{first_row}:F{last_row}, "ADEUDA FINAL")', False),
            ("% Cursadas Completadas", f'=COUNTIF(B{first_row}:B{last_row}, "SI")/{total_materias}', True),
            ("% Carrera Aprobada (Finales)", f'=COUNTIF(C{first_row}:C{last_row}, "SI")/{total_materias}', True),
        ]

        for s_idx, (label, formula_val, is_pct) in enumerate(stat_items, 2):
            ws.cell(row=s_idx, column=9, value=label).font = font_bold
            ws.cell(row=s_idx, column=9).fill = light_gray_fill
            ws.cell(row=s_idx, column=9).border = thin_border

            v_cell = ws.cell(row=s_idx, column=10, value=formula_val)
            v_cell.font = font_indicator
            v_cell.alignment = Alignment(horizontal="center", vertical="center")
            v_cell.border = thin_border
            v_cell.fill = green_soft if is_pct else light_blue_fill
            if is_pct:
                v_cell.number_format = '0.0%'

            ws.row_dimensions[s_idx].height = 22

        # 4. Cuadro de Instrucciones
        inst_row = 9
        ws.cell(row=inst_row, column=9, value="INSTRUCCIONES DE USO").font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        ws.cell(row=inst_row, column=9).fill = gold_fill
        ws.cell(row=inst_row, column=9).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=inst_row, start_column=9, end_row=inst_row, end_column=10)

        instrucciones = [
            "1. Hace clic en la celda y elegi 'SI' o 'NO' desde el menu desplegable.",
            "2. Columna B: Marca 'SI' si tenes la cursada regular aprobada.",
            "3. Columna C: Marca 'SI' si aprobaste el final o promocionaste.",
            "4. Las columnas E, F y G se actualizan automaticamente en tiempo real.",
            "5. El panel lateral calcula tu avance y porcentaje de carrera en vivo.",
            "6. Podes subir el archivo a tu Google Drive personal para editarlo online."
        ]

        for idx_i, inst in enumerate(instrucciones, inst_row + 1):
            cell = ws.cell(row=idx_i, column=9, value=inst)
            cell.font = Font(name="Arial", size=8.5, italic=True)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.merge_cells(start_row=idx_i, start_column=9, end_row=idx_i, end_column=10)
            ws.row_dimensions[idx_i].height = 20

    # 5. Ajustar Ancho de Columnas
    ws.column_dimensions["A"].width = 54
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 4
    ws.column_dimensions["I"].width = 30
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 4

    out_path = os.path.join(OUTPUT_DIR, filename)
    wb.save(out_path)
    print(f"Generado: {filename} ({total_materias} materias)")

def main():
    with open(os.path.join(BASE_DIR, "js", "upload_form.js"), "r", encoding="utf-8") as f:
        text = f.read()

    idx = text.find("const CURRICULUM_DATA = ")
    end_idx = text.find(";\n\ndocument.addEventListener")
    json_str = text[idx + len("const CURRICULUM_DATA = "):end_idx]
    CURRICULUM_DATA = json.loads(json_str)

    generated_files = set()

    for career_name, years_data in CURRICULUM_DATA.items():
        filename = CAREER_FILENAMES.get(career_name)
        if not filename:
            clean_name = career_name.replace("Profesorado de ", "").replace(" ", "_")
            filename = f"Mapa_de_Carrera_{clean_name}.xlsx"

        if filename not in generated_files:
            build_career_excel(career_name, years_data, filename)
            generated_files.add(filename)

    # Plantilla General
    plantilla_data = {
        "1° Año": ["Pedagogía", "Didáctica General", "Lectura, Escritura y Oralidad I (LEO 1)", "Práctica Docente I"],
        "2° Año": ["Psicología Educacional", "Sociología de la Educación", "Sujetos de la Educación", "Práctica Docente II"],
        "3° Año": ["Historia Social de la Educación", "Residencia I", "Espacio de Definición Institucional", "Práctica Docente III"],
        "4° Año": ["Residencia II / Taller de Cierre", "Ética y Trabajo Docente", "Investigación Educativa", "Práctica Docente IV"]
    }
    build_career_excel("General (Plantilla Multicarrera)", plantilla_data, "Mapa_de_Carrera_General_Plantilla.xlsx")
    print(f"\nTotal de mapas generados con exito: {len(generated_files) + 1}")

if __name__ == "__main__":
    main()
