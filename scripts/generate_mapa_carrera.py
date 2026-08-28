"""
Script para generar los archivos oficiales de "Mapa de Carrera Personal" (.xlsx)
con la estética, marca de agua e identidad de:
La Caravana + Estudiantes Independientes (Lista 90) - AltilloJVG

Incluye menú desplegable nativo (Dropdown) con opciones 'SI' / 'NO'
para que las y los estudiantes no tengan que escribir nada a mano.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(BASE_DIR, "data", "mapas")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 1. Definición de materias para Profesorado de Filosofía
FILOSOFIA_DATA = [
    ("Primer Año", [
        ("Lógica 1 y Teoría de la Argumentación", []),
        ("Introducción a la Filosofía", []),
        ("Psicología Educacional", []),
        ("Psicología del Desarrollo", []),
        ("Sociología", []),
        ("Pedagogía", []),
        ("LEO 1 (Lectura y Escritura Académica)", []),
        ("Trabajo de Campo 1", []),
        ("Educación Sexual Integral (ESI)", []),
        ("Lengua Extranjera", [])
    ]),
    ("Segundo Año", [
        ("Didáctica General", ["Pedagogía"]),
        ("Historia de la Educación Argentina", ["Pedagogía", "Sociología"]),
        ("Historia de la Filosofía Antigua", ["Introducción a la Filosofía"]),
        ("Filosofía Argentina y Latinoamericana", ["Introducción a la Filosofía"]),
        ("Historia del Arte", []),
        ("Metodología de la Investigación", ["Lógica 1 y Teoría de la Argumentación"]),
        ("Nuevas Tecnologías", []),
        ("LEO 2", ["LEO 1 (Lectura y Escritura Académica)"]),
        ("Trabajo de Campo 2", ["Trabajo de Campo 1"])
    ]),
    ("Tercer Año", [
        ("Lógica 2", ["Lógica 1 y Teoría de la Argumentación"]),
        ("Filosofía de la Lógica", ["Lógica 1 y Teoría de la Argumentación"]),
        ("Historia de la Filosofía Medieval", ["Historia de la Filosofía Antigua"]),
        ("Metafísica", ["Historia de la Filosofía Antigua"]),
        ("Historia de la Ciencia", ["Metodología de la Investigación"]),
        ("Trabajo de Campo 3", ["Trabajo de Campo 2"]),
        ("Taller de Didáctica de la Filosofía y Prod. Materiales", ["Didáctica General"]),
        ("Sistema y Política Educativa", ["Historia de la Educación Argentina"]),
        ("Taller Filosofía y Educación", ["Historia de la Filosofía Antigua"]),
        ("Teoría del Conocimiento", ["Historia de la Filosofía Antigua"]),
        ("Ética", ["Introducción a la Filosofía"])
    ]),
    ("Cuarto Año", [
        ("Historia de la Filosofía Moderna", ["Historia de la Filosofía Medieval"]),
        ("Filosofía Política", ["Filosofía Argentina y Latinoamericana", "Ética"]),
        ("Filosofía del Lenguaje", ["Filosofía de la Lógica"]),
        ("Antropología Filosófica", ["Historia de la Filosofía Antigua"]),
        ("Estética", ["Historia del Arte"]),
        ("Práctica Docente 1", ["Taller de Didáctica de la Filosofía y Prod. Materiales", "Trabajo de Campo 3"])
    ]),
    ("Quinto Año", [
        ("Historia de la Filosofía Contemporánea", ["Historia de la Filosofía Moderna"]),
        ("Filosofía del Derecho", ["Filosofía Política"]),
        ("Filosofía de la Religión", ["Historia de la Filosofía Medieval"]),
        ("Filosofía de la Historia", ["Historia de la Filosofía Moderna"]),
        ("Filosofía de la Ciencia", ["Historia de la Ciencia", "Teoría del Conocimiento"]),
        ("Práctica Docente 2 y Residencia", ["Práctica Docente 1"])
    ])
]

def build_mapa_filosofia():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapa de Carrera"

    # Hoja de Opciones para Menú Desplegable (Hidden)
    ws_opt = wb.create_sheet(title="Opciones")
    ws_opt["A1"] = "SI"
    ws_opt["A2"] = "NO"
    ws_opt.sheet_state = 'hidden'

    # Paleta de Colores
    navy_fill = PatternFill(start_color="0B2545", end_color="0B2545", fill_type="solid")
    cyan_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    gold_fill = PatternFill(start_color="D99B26", end_color="D99B26", fill_type="solid")
    light_blue_fill = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
    light_gray_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    section_fill = PatternFill(start_color="134074", end_color="134074", fill_type="solid")
    green_soft = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

    font_title = Font(name="Arial", size=15, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Arial", size=10, italic=True, color="BAE6FD")
    font_section = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_regular = Font(name="Arial", size=10)
    font_bold = Font(name="Arial", size=10, bold=True)
    font_indicator = Font(name="Arial", size=11, bold=True, color="0B2545")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # 1. Encabezado Institucional
    ws.merge_cells("A1:G1")
    ws["A1"] = "MAPA DE CARRERA PERSONAL | ALTILLO JVG"
    ws["A1"].font = font_title
    ws["A1"].fill = navy_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:G2")
    ws["A2"] = "Desarrollado por La Caravana + Estudiantes Independientes (Lista 90) - Joaquín V. González"
    ws["A2"].font = font_subtitle
    ws["A2"].fill = section_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # 2. Encabezados de Columnas
    headers = [
        ("A3", "Asignatura / Materia", 38),
        ("B3", "¿Cursada Regular? (Elegir SI/NO)", 22),
        ("C3", "¿Final Aprobado? (Elegir SI/NO)", 22),
        ("D3", "Correlativas para Cursar", 32),
        ("E3", "¿Podés Cursar?", 18),
        ("F3", "¿Podés Rendir Final?", 20),
        ("G3", "Estado Trayectoria", 22)
    ]

    for cell_id, text, col_width in headers:
        cell = ws[cell_id]
        cell.value = text
        cell.font = font_header
        cell.fill = cyan_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col_letter = cell_id[0]
        ws.column_dimensions[col_letter].width = col_width
    ws.row_dimensions[3].height = 28

    current_row = 4
    materia_row_map = {}
    subject_rows = []

    # 3. Llenado de Materias por Año
    for anio_label, materias in FILOSOFIA_DATA:
        # Fila de Año
        ws.merge_cells(f"A{current_row}:G{current_row}")
        section_cell = ws[f"A{current_row}"]
        section_cell.value = f"📌 {anio_label.upper()}"
        section_cell.font = font_section
        section_cell.fill = section_fill
        section_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        for mat_name, correlativas in materias:
            materia_row_map[mat_name] = current_row
            r = current_row
            subject_rows.append(r)
            row_fill = light_gray_fill if (r % 2 == 0) else PatternFill(fill_type=None)
            
            # Col A: Nombre
            ws[f"A{r}"] = mat_name
            ws[f"A{r}"].font = font_regular
            ws[f"A{r}"].border = thin_border
            ws[f"A{r}"].alignment = Alignment(vertical="center")
            if row_fill.fill_type: ws[f"A{r}"].fill = row_fill

            # Col B: Cursada Aprobada (Por defecto NO)
            ws[f"B{r}"] = "NO"
            ws[f"B{r}"].font = font_bold
            ws[f"B{r}"].border = thin_border
            ws[f"B{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # Col C: Final Aprobado (Por defecto NO)
            ws[f"C{r}"] = "NO"
            ws[f"C{r}"].font = font_bold
            ws[f"C{r}"].border = thin_border
            ws[f"C{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # Col D: Texto de Correlativas
            ws[f"D{r}"] = ", ".join(correlativas) if correlativas else "Sin correlativas previas"
            ws[f"D{r}"].font = Font(name="Arial", size=9, italic=(not correlativas), color="64748B" if not correlativas else "0F172A")
            ws[f"D{r}"].border = thin_border
            ws[f"D{r}"].alignment = Alignment(vertical="center")
            if row_fill.fill_type: ws[f"D{r}"].fill = row_fill

            # Condiciones universales para SI / SÍ / TRUE
            is_reg = f'OR(B{r}="SI",B{r}="SÍ",B{r}=TRUE)'
            is_fin = f'OR(C{r}="SI",C{r}="SÍ",C{r}=TRUE)'

            # Col E: ¿Podés Cursar? (Fórmula condicional)
            if not correlativas:
                ws[f"E{r}"] = f'=IF({is_reg},"CURSADA REGULAR","¡HABILITADA!")'
            else:
                # Comprobar que todas las correlativas tengan cursada regular en Columna B
                conds = [f'OR(B{materia_row_map[c]}="SI",B{materia_row_map[c]}="SÍ",B{materia_row_map[c]}=TRUE)' for c in correlativas if c in materia_row_map]
                if conds:
                    and_clause = f"AND({','.join(conds)})"
                    ws[f"E{r}"] = f'=IF({is_reg},"CURSADA REGULAR",IF({and_clause},"¡HABILITADA!","Faltan correlat."))'
                else:
                    ws[f"E{r}"] = f'=IF({is_reg},"CURSADA REGULAR","¡HABILITADA!")'
            
            ws[f"E{r}"].font = font_bold
            ws[f"E{r}"].border = thin_border
            ws[f"E{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # Col F: ¿Podés Rendir Final?
            if not correlativas:
                ws[f"F{r}"] = f'=IF({is_fin},"APROBADA",IF({is_reg},"¡LISTO P/ FINAL!","Falta cursada"))'
            else:
                conds_final = [f'OR(C{materia_row_map[c]}="SI",C{materia_row_map[c]}="SÍ",C{materia_row_map[c]}=TRUE)' for c in correlativas if c in materia_row_map]
                if conds_final:
                    and_final = f"AND({','.join(conds_final)})"
                    ws[f"F{r}"] = f'=IF({is_fin},"APROBADA",IF(AND({is_reg},{and_final}),"¡LISTO P/ FINAL!","Faltan finales prev."))'
                else:
                    ws[f"F{r}"] = f'=IF({is_fin},"APROBADA",IF({is_reg},"¡LISTO P/ FINAL!","Falta cursada"))'

            ws[f"F{r}"].font = font_bold
            ws[f"F{r}"].border = thin_border
            ws[f"F{r}"].alignment = Alignment(horizontal="center", vertical="center")

            # Col G: Estado General
            ws[f"G{r}"] = f'=IF({is_fin},"✅ Materia Aprobada",IF({is_reg},"⏳ Regular (Adeuda Final)","⚪ Pendiente"))'
            ws[f"G{r}"].font = font_bold
            ws[f"G{r}"].border = thin_border
            ws[f"G{r}"].alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[r].height = 20
            current_row += 1

    last_data_row = current_row - 1
    total_materias = len(materia_row_map)

    # 4. Menú Desplegable Nativo (Dropdown) con SI / NO para cada celda
    dv = DataValidation(type="list", formula1="=Opciones!$A$1:$A$2", allow_blank=True)
    dv.error ='Elegí una opción de la lista'
    dv.errorTitle = 'Opción no válida'
    dv.prompt = 'Elegí SI o NO'
    dv.promptTitle = 'Estado de Materia'
    ws.add_data_validation(dv)

    for r in subject_rows:
        dv.add(f"B{r}")
        dv.add(f"C{r}")

    # 5. Panel Lateral de Estadísticas y Avance
    ws.column_dimensions["I"].width = 30
    ws.column_dimensions["J"].width = 18

    ws.merge_cells("I3:J3")
    ws["I3"] = "📊 RESUMEN DE AVANCE"
    ws["I3"].font = font_header
    ws["I3"].fill = navy_fill
    ws["I3"].alignment = Alignment(horizontal="center", vertical="center")

    count_reg_formula = f'=COUNTIF(B4:B{last_data_row}, "S*") + COUNTIF(B4:B{last_data_row}, TRUE)'
    count_fin_formula = f'=COUNTIF(C4:C{last_data_row}, "S*") + COUNTIF(C4:C{last_data_row}, TRUE)'

    summary_rows = [
        ("Total Materias del Plan", total_materias, False),
        ("Cursadas Regulares", count_reg_formula, True),
        ("Finales Aprobados", count_fin_formula, True),
        ("Finales Adeudados", "=J5-J6", True),
        ("% Cursadas Completadas", f'=J5/{total_materias}', True),
        ("% Carrera Aprobada (Finales)", f'=J6/{total_materias}', True)
    ]

    r_idx = 4
    for label, val, is_formula in summary_rows:
        ws[f"I{r_idx}"] = label
        ws[f"I{r_idx}"].font = font_bold
        ws[f"I{r_idx}"].border = thin_border
        ws[f"I{r_idx}"].fill = light_blue_fill
        ws[f"I{r_idx}"].alignment = Alignment(vertical="center")

        ws[f"J{r_idx}"] = val
        ws[f"J{r_idx}"].font = font_indicator
        ws[f"J{r_idx}"].border = thin_border
        ws[f"J{r_idx}"].alignment = Alignment(horizontal="center", vertical="center")
        
        if "%" in label:
            ws[f"J{r_idx}"].number_format = '0.0%'
            ws[f"J{r_idx}"].fill = green_soft

        ws.row_dimensions[r_idx].height = 24
        r_idx += 1

    # Instrucciones de uso en el Excel
    r_idx += 1
    ws.merge_cells(f"I{r_idx}:J{r_idx}")
    ws[f"I{r_idx}"] = "💡 INSTRUCCIONES DE USO"
    ws[f"I{r_idx}"].font = font_header
    ws[f"I{r_idx}"].fill = gold_fill
    ws[f"I{r_idx}"].alignment = Alignment(horizontal="center", vertical="center")
    r_idx += 1

    instructions = [
        "1. Hacé clic en la celda y elegí 'SI' o 'NO' desde el menú desplegable.",
        "2. Columna B: Marcá 'SI' si tenés la cursada regular aprobada.",
        "3. Columna C: Marcá 'SI' si aprobaste el final o promocionaste.",
        "4. Las columnas E y F te informan automáticamente qué podés cursar o rendir.",
        "5. El panel lateral calcula tu avance y porcentaje de carrera en tiempo real.",
        "6. Guardá una copia en tu Google Drive personal para conservar tus progresos."
    ]
    for inst in instructions:
        ws.merge_cells(f"I{r_idx}:J{r_idx}")
        ws[f"I{r_idx}"] = inst
        ws[f"I{r_idx}"].font = font_regular
        ws[f"I{r_idx}"].alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r_idx].height = 26
        r_idx += 1

    out_file = os.path.join(OUTPUT_DIR, "Mapa_de_Carrera_Filosofia.xlsx")
    wb.save(out_file)
    print(f"[+] Archivo regenerado con menú desplegable nativo: {out_file}")

if __name__ == "__main__":
    build_mapa_filosofia()
