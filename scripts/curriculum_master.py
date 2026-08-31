"""
curriculum_master.py
Planes de estudio canónicos oficiales y curados meticulosamente
para los 18 profesorados del Instituto Superior del Profesorado "Joaquín V. González".
Libre de duplicados, errores tipográficos o sufijos de comisión.
"""
import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

MASTER_CURRICULUM = {
    "Profesorado de Matemática": {
        "1° Año": [
            "Álgebra I",
            "Análisis Matemático I",
            "Geometría I",
            "Introducción a la Matemática Superior",
            "Elementos Básicos de Matemática",
            "Pedagogía",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Trabajo de Campo I"
        ],
        "2° Año": [
            "Álgebra II",
            "Análisis Matemático II",
            "Geometría II",
            "Física General",
            "Didáctica General",
            "Psicología Educacional",
            "Sujetos de la Educación",
            "Trabajo de Campo II"
        ],
        "3° Año": [
            "Álgebra III",
            "Análisis Matemático III",
            "Geometría III",
            "Probabilidad y Estadística",
            "Didáctica de la Matemática I",
            "Historia Social de la Educación",
            "Derechos Humanos, Sociedad y Estado",
            "Construcción de la Práctica Docente I"
        ],
        "4° Año": [
            "Fundamentos de la Matemática",
            "Ecuaciones Diferenciales",
            "Didáctica de la Matemática II",
            "Historia de la Matemática",
            "Tecnologías Digitales en la Enseñanza de la Matemática",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente",
            "Residencia y Construcción de la Práctica Docente II"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Especialización en Matemática",
            "Epistemología de la Matemática",
            "Taller de Investigación Educativa en Matemática"
        ]
    },
    "Profesorado de Filosofía": {
        "1° Año": [
            "Introducción a la Filosofía",
            "Lógica I",
            "Filosofía Antigua",
            "Historia de la Filosofía Antigua",
            "Pedagogía",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Trabajo de Campo I"
        ],
        "2° Año": [
            "Filosofía Medieval",
            "Gnoseología",
            "Antropología Filosófica",
            "Lógica II",
            "Didáctica General",
            "Psicología Educacional",
            "Sujetos de la Educación",
            "Trabajo de Campo II"
        ],
        "3° Año": [
            "Filosofía Moderna",
            "Metafísica",
            "Ética",
            "Filosofía Política",
            "Didáctica de la Filosofía",
            "Historia Social de la Educación",
            "Derechos Humanos, Sociedad y Estado",
            "Construcción de la Práctica Docente I"
        ],
        "4° Año": [
            "Filosofía Contemporánea",
            "Filosofía Latinoamericana y Argentina",
            "Estética",
            "Filosofía de las Ciencias y Epistemología",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente",
            "Residencia y Construcción de la Práctica Docente II"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Textos Filosóficos I",
            "Seminario de Textos Filosóficos II",
            "Taller de Investigación en Filosofía"
        ]
    },
    "Profesorado de Historia": {
        "1° Año": [
            "Introducción a la Historia",
            "Historia Antigua de Oriente y Grecia",
            "Historia Antigua de Roma",
            "Historia de América I (Precolombina e Indígena)",
            "Pedagogía",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Trabajo de Campo I"
        ],
        "2° Año": [
            "Historia Medieval",
            "Historia Moderna",
            "Historia de América II (Colonial)",
            "Historia Argentina I (Colonial e Independencia)",
            "Didáctica General",
            "Psicología Educacional",
            "Sujetos de la Educación",
            "Trabajo de Campo II"
        ],
        "3° Año": [
            "Historia Contemporánea del Siglo XIX",
            "Historia de América III (Siglo XIX)",
            "Historia Argentina II (Siglo XIX)",
            "Didáctica de la Historia I",
            "Historia Social de la Educación",
            "Derechos Humanos, Sociedad y Estado",
            "Construcción de la Práctica Docente I"
        ],
        "4° Año": [
            "Historia Contemporánea del Siglo XX",
            "Historia de América IV (Siglo XX y Actual)",
            "Historia Argentina III (Siglo XX y Actual)",
            "Historiografía",
            "Didáctica de la Historia II",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente",
            "Residencia y Construcción de la Práctica Docente II"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Historia Argentina y Latinoamericana",
            "Seminario de Historia Mundial Contemporánea",
            "Taller de Investigación Histórica"
        ]
    },
    "Profesorado de Lengua y Literatura": {
        "1° Año": [
            "Lingüística General I",
            "Gramática Española I",
            "Teoría y Análisis Literario I",
            "Literatura Española I (Medieval y Siglo de Oro)",
            "Pedagogía",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Trabajo de Campo I"
        ],
        "2° Año": [
            "Lingüística General II",
            "Gramática Española II",
            "Teoría y Análisis Literario II",
            "Literatura Española II",
            "Literatura Latinoamericana I",
            "Didáctica General",
            "Psicología Educacional",
            "Sujetos de la Educación",
            "Trabajo de Campo II"
        ],
        "3° Año": [
            "Historia de la Lengua Española",
            "Literatura Latinoamericana II",
            "Literatura Argentina I",
            "Didáctica de la Lengua y la Literatura I",
            "Historia Social de la Educación",
            "Derechos Humanos, Sociedad y Estado",
            "Construcción de la Práctica Docente I"
        ],
        "4° Año": [
            "Literatura Argentina II",
            "Literaturas Extranjeras Clásicas y Modernas",
            "Semiótica y Análisis del Discurso",
            "Didáctica de la Lengua y la Literatura II",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente",
            "Residencia y Construcción de la Práctica Docente II"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Literatura Argentina y Latinoamericana",
            "Seminario de Lingüística Aplicada",
            "Taller de Escritura Creativa y Académica"
        ]
    },
    "Profesorado de Inglés": {
        "1° Año": [
            "Lengua Inglesa I",
            "Fonética y Fonología Inglesa I",
            "Gramática Inglesa I",
            "Dicción y Práctica de Laboratorio I",
            "Geografía y Cultura de los Pueblos de Habla Inglesa I",
            "Pedagogía",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Construcción de la Práctica Docente en Nivel Inicial y Primario I"
        ],
        "2° Año": [
            "Lengua Inglesa II",
            "Fonética y Fonología Inglesa II",
            "Gramática Inglesa II",
            "Literatura en Lengua Inglesa I",
            "Geografía y Cultura de los Pueblos de Habla Inglesa II",
            "Didáctica General",
            "Psicología Educacional",
            "Sujetos de la Educación",
            "Construcción de la Práctica Docente en Nivel Inicial y Primario II y Residencia"
        ],
        "3° Año": [
            "Lengua Inglesa III",
            "Fonología y Fonética Inglesa III",
            "Gramática Inglesa III",
            "Lingüística General y Aplicada",
            "Literatura en Lengua Inglesa II",
            "Historia de la Lengua Inglesa",
            "Didáctica Especial y Construcción de la Práctica Docente en Nivel Medio I",
            "Historia Social de la Educación",
            "Derechos Humanos, Sociedad y Estado"
        ],
        "4° Año": [
            "Lengua Inglesa IV",
            "Análisis del Discurso en Lengua Inglesa",
            "Literatura en Lengua Inglesa III y Contemporánea",
            "Estudios Culturales Anglófonos",
            "Didáctica Especial y Residencia en Nivel Medio y Superior",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Literatura en Lengua Inglesa",
            "Seminario de Lingüística y Adquisición del Lenguaje",
            "Taller de Traducción y Práctica Profesional"
        ]
    },
    "Profesorado de Biología": {
        "1° Año": [
            "Introducción a la Biología y Sistemas Vivientes",
            "Biología Celular y Molecular",
            "Diversidad Vegetal I",
            "Química General e Inorgánica",
            "Matemática y Estadística Aplicada",
            "Pedagogía",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Trabajo de Campo I"
        ],
        "2° Año": [
            "Diversidad Vegetal II",
            "Diversidad Animal I (Invertebrados)",
            "Anatomía y Fisiología Humana",
            "Química Orgánica y Biológica",
            "Didáctica General",
            "Psicología Educacional",
            "Sujetos de la Educación",
            "Trabajo de Campo II"
        ],
        "3° Año": [
            "Diversidad Animal II (Vertebrados)",
            "Genética General y Molecular",
            "Ecología General y de Poblaciones",
            "Didáctica de la Biología I",
            "Ciencias de la Tierra y Dinámica Terrestre",
            "Historia Social de la Educación",
            "Derechos Humanos, Sociedad y Estado",
            "Construcción de la Práctica Docente I"
        ],
        "4° Año": [
            "Evolución y Paleontología",
            "Fisiología y Morfología Comparada",
            "Educación para la Salud y el Ambiente",
            "Didáctica de la Biología II",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente",
            "Residencia y Construcción de la Práctica Docente II"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Biotecnología y Sociedad",
            "Seminario de Epistemología de la Biología",
            "Taller de Investigación en Ciencias Naturales"
        ]
    },
    "Profesorado de Física": {
        "1° Año": [
            "Física General I (Mecánica y Cinemática)",
            "Álgebra y Geometría Analítica I",
            "Análisis Matemático I",
            "Laboratorio de Física I",
            "Pedagogía",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Trabajo de Campo I"
        ],
        "2° Año": [
            "Física General II (Calor, Termodinámica y Fluidos)",
            "Física General III (Óptica y Ondas)",
            "Análisis Matemático II",
            "Química General",
            "Laboratorio de Física II",
            "Didáctica General",
            "Psicología Educacional",
            "Sujetos de la Educación",
            "Trabajo de Campo II"
        ],
        "3° Año": [
            "Física General IV (Electromagnetismo)",
            "Métodos Matemáticos de la Física",
            "Mecánica Teórica",
            "Didáctica de la Física I",
            "Laboratorio de Física III",
            "Historia Social de la Educación",
            "Derechos Humanos, Sociedad y Estado",
            "Construcción de la Práctica Docente I"
        ],
        "4° Año": [
            "Introducción a la Física Moderna y Cuántica",
            "Termodinámica Estadística",
            "Astronomía y Astrofísica Básica",
            "Didáctica de la Física II",
            "Historia y Epistemología de la Física",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente",
            "Residencia y Construcción de la Práctica Docente II"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Física Contemporánea",
            "Taller de Experimentación y Modelización",
            "Taller de Investigación en Enseñanza de la Física"
        ]
    },
    "Profesorado de Química": {
        "1° Año": [
            "Química General e Inorgánica I",
            "Laboratorio de Química General",
            "Matemática I (Cálculo y Álgebra)",
            "Física I",
            "Pedagogía",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Trabajo de Campo I"
        ],
        "2° Año": [
            "Química General e Inorgánica II",
            "Química Orgánica I",
            "Química Analítica Cualitativa",
            "Matemática II",
            "Física II",
            "Didáctica General",
            "Psicología Educacional",
            "Sujetos de la Educación",
            "Trabajo de Campo II"
        ],
        "3° Año": [
            "Química Orgánica II",
            "Química Analítica Cuantitativa e Instrumental",
            "Fisicoquímica I",
            "Didáctica de la Química I",
            "Biología Celular y Bioquímica",
            "Historia Social de la Educación",
            "Derechos Humanos, Sociedad y Estado",
            "Construcción de la Práctica Docente I"
        ],
        "4° Año": [
            "Fisicoquímica II",
            "Química Industrial y Ambiental",
            "Bromatología y Toxicología",
            "Didáctica de la Química II",
            "Historia y Epistemología de la Química",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente",
            "Residencia y Construcción de la Práctica Docente II"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Química Aplicada",
            "Taller de Seguridad e Higiene en Laboratorio",
            "Taller de Investigación en Enseñanza de la Química"
        ]
    },
    "Profesorado de Geografía": {
        "1° Año": [
            "Introducción a la Geografía",
            "Cartografía y Sistemas de Información Geográfica (SIG)",
            "Geografía Física I (Geomorfología y Geología)",
            "Geografía Humana y Social I",
            "Pedagogía",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Trabajo de Campo I"
        ],
        "2° Año": [
            "Geografía Física II (Climatología, Hidrología y Biogeografía)",
            "Geografía Económica y Urbana",
            "Geografía de América Latina",
            "Geografía Rural y de la Población",
            "Didáctica General",
            "Psicología Educacional",
            "Sujetos de la Educación",
            "Trabajo de Campo II"
        ],
        "3° Año": [
            "Geografía de la República Argentina I (Bases Físicas y Ambientales)",
            "Geografía de la República Argentina II (Socioeconómica y Regional)",
            "Geopolítica y Geografía Política",
            "Didáctica de la Geografía I",
            "Historia Social de la Educación",
            "Derechos Humanos, Sociedad y Estado",
            "Construcción de la Práctica Docente I"
        ],
        "4° Año": [
            "Geografía Mundial Contemporánea (Problemas Globales)",
            "Planificación Territorial y Problemáticas Ambientales",
            "Epistemología e Historia del Pensamiento Geográfico",
            "Didáctica de la Geografía II",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente",
            "Residencia y Construcción de la Práctica Docente II"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Geografía Urbana y Metropolización",
            "Seminario de Recursos Naturales y Desarrollo Sustentable",
            "Taller de Investigación Geográfica"
        ]
    },
    "Profesorado de Informática": {
        "1° Año": [
            "Algoritmos y Estructuras de Datos I",
            "Arquitectura y Organización de Computadoras",
            "Fundamentos de Informática y Sistemas Digitales",
            "Matemática Discreta y Álgebra Lineal",
            "Pedagogía",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Trabajo de Campo I"
        ],
        "2° Año": [
            "Algoritmos y Estructuras de Datos II",
            "Sistemas Operativos y Redes de Comunicación",
            "Bases de Datos I",
            "Programación Orientada a Objetos",
            "Didáctica General",
            "Psicología Educacional",
            "Sujetos de la Educación",
            "Trabajo de Campo II"
        ],
        "3° Año": [
            "Ingeniería de Software y Desarrollo Web",
            "Bases de Datos II y Administración de Servidores",
            "Informática Educativa I y Entornos Virtuales",
            "Didáctica de la Informática I",
            "Seguridad Informática y Ética Digital",
            "Historia Social de la Educación",
            "Derechos Humanos, Sociedad y Estado",
            "Construcción de la Práctica Docente I"
        ],
        "4° Año": [
            "Tecnologías Emergentes, IA y Ciencia de Datos",
            "Informática Educativa II y Robótica Pedagógica",
            "Discursos y Culturas Digitales",
            "Didáctica de la Informática II",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente",
            "Residencia y Construcción de la Práctica Docente II"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Redes Avanzadas y Cloud Computing",
            "Seminario de Accesibilidad y Tecnologías Inclusivas",
            "Taller de Proyecto Final de Informática"
        ]
    },
    "Profesorado de Psicología": {
        "1° Año": [
            "Psicología General",
            "Antropología Social y Cultural",
            "Lógica y Epistemología General",
            "Biología y Neurociencias del Comportamiento",
            "Procesos Colectivos y Problemas Sociales",
            "Pedagogía",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Trabajo de Campo I"
        ],
        "2° Año": [
            "Psicología del Desarrollo y Ciclos Vitales I (Niñez y Adolescencia)",
            "Psicoanálisis I (Teoría Freudiana)",
            "Psicología Social y Comunitaria",
            "Didáctica General",
            "Psicología Educacional",
            "Sujetos de la Educación",
            "Estado, Sociedad y Derechos Humanos",
            "Trabajo de Campo II"
        ],
        "3° Año": [
            "Psicología del Desarrollo y Ciclos Vitales II (Adultez y Vejez)",
            "Psicoanálisis II (Teorías Posfreudianas)",
            "Teoría de las Organizaciones y Análisis Institucional",
            "Psicopatología Infanto-Juvenil y del Adulto",
            "Técnicas de Exploración y Evaluación Psicológica",
            "Didáctica de la Enseñanza de la Psicología I",
            "Historia Social de la Educación",
            "Construcción de la Práctica Docente I"
        ],
        "4° Año": [
            "Historia de la Psicología",
            "Psicología Institucional y Forense",
            "Psicopedagogía y Orientación Vocacional",
            "Didáctica de la Enseñanza de la Psicología II",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente",
            "Residencia y Construcción de la Práctica Docente II"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Psicología Comunitaria e Intervención Social",
            "Seminario de Epistemología Psicológica",
            "Taller de Investigación en Psicología"
        ]
    },
    "Profesorado de Ciencias de la Educación": {
        "1° Año": [
            "Pedagogía General",
            "Filosofía General",
            "Antropología y Educación",
            "Psicología General y del Desarrollo",
            "Sociología General",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Trabajo de Campo I"
        ],
        "2° Año": [
            "Didáctica General",
            "Historia de la Educación Argentina y Latinoamericana",
            "Sociología de la Educación",
            "Psicología de la Educación",
            "Instituciones Educativas y Análisis Organizacional",
            "Filosofía de la Educación",
            "Estado, Sociedad y Derechos Humanos",
            "Trabajo de Campo II"
        ],
        "3° Año": [
            "Teoría y Diseño Curricular",
            "Política y Legislación Educativa",
            "Metodología de la Investigación Educativa I (Cualitativa)",
            "Didáctica de Nivel Inicial, Primario y Secundario",
            "Historia Social de la Educación",
            "Problemas Didácticos Contemporáneos",
            "Construcción de la Práctica Docente I"
        ],
        "4° Año": [
            "Planeamiento y Gestión Educativa",
            "Metodología de la Investigación Educativa II (Cuantitativa y Mixta)",
            "Evaluación de Proyectos Educativos",
            "Educación No Formal y Pedagogía Social",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente",
            "Residencia y Construcción de la Práctica Docente II"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Políticas Públicas Educativas",
            "Seminario de Educación Superior y Universitaria",
            "Taller de Trabajo Final / Tesis de Licenciatura"
        ]
    },
    "Profesorado de Ciencias Jurídicas": {
        "1° Año": [
            "Introducción al Derecho y Teoría General",
            "Derecho Político y Teoría del Estado",
            "Historia de las Ideas e Instituciones Políticas",
            "Sociología Jurídica",
            "Pedagogía",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Trabajo de Campo I"
        ],
        "2° Año": [
            "Derecho Constitucional y Garantías",
            "Derechos Humanos y Derecho Internacional de los DD.HH.",
            "Derecho Privado I (Parte General y Personas)",
            "Derecho Penal I (Parte General)",
            "Didáctica General",
            "Psicología Educacional",
            "Sujetos de la Educación",
            "Trabajo de Campo II"
        ],
        "3° Año": [
            "Derecho Privado II (Obligaciones y Contratos)",
            "Derecho Administrativo",
            "Derecho del Trabajo y Relaciones Laborales",
            "Didáctica de las Ciencias Jurídicas I",
            "Elementos de Derecho Procesal",
            "Historia Social de la Educación",
            "Construcción de la Práctica Docente I"
        ],
        "4° Año": [
            "Derecho Internacional Público y de la Integración",
            "Finanzas Públicas y Derecho Tributario",
            "Derecho Ambiental y de los Recursos Naturales",
            "Didáctica de las Ciencias Jurídicas II",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente",
            "Residencia y Construcción de la Práctica Docente II"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Derecho Constitucional Profundizado",
            "Seminario de Género y Sistema de Justicia",
            "Taller de Práctica Forense y Litigación Educativa"
        ]
    },
    "Profesorado de Ciencia Política": {
        "1° Año": [
            "Introducción a la Ciencia Política",
            "Teoría Política Clásica y Medieval",
            "Historia Política Argentina I",
            "Sociología Política",
            "Pedagogía",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Trabajo de Campo I"
        ],
        "2° Año": [
            "Teoría Política Moderna",
            "Historia Política Argentina II",
            "Sistemas Políticos Comparados",
            "Opinión Pública y Comunicación Política",
            "Didáctica General",
            "Psicología Educacional",
            "Sujetos de la Educación",
            "Trabajo de Campo II"
        ],
        "3° Año": [
            "Teoría Política Contemporánea",
            "Relaciones Internacionales y Política Exterior",
            "Políticas Públicas y Administración del Estado",
            "Didáctica de la Ciencia Política I",
            "Historia Social de la Educación",
            "Derechos Humanos, Sociedad y Estado",
            "Construcción de la Práctica Docente I"
        ],
        "4° Año": [
            "Partidos Políticos y Sistemas Electorales",
            "Filosofía Política Latinoamericana",
            "Metodología del Análisis Político",
            "Didáctica de la Ciencia Política II",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente",
            "Residencia y Construcción de la Práctica Docente II"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Política Latinoamericana Contemporánea",
            "Seminario de Movimientos Sociales y Ciudadanía",
            "Taller de Investigación en Ciencia Política"
        ]
    },
    "Profesorado de Economía": {
        "1° Año": [
            "Principios de Economía y Microeconomía I",
            "Matemática para Economistas I (Álgebra y Cálculo)",
            "Historia del Pensamiento Económico I",
            "Contabilidad General y Social",
            "Pedagogía",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Trabajo de Campo I"
        ],
        "2° Año": [
            "Macroeconomía I",
            "Microeconomía II",
            "Matemática para Economistas II",
            "Estadística Económica",
            "Historia Económica Argentina y Mundial",
            "Didáctica General",
            "Psicología Educacional",
            "Sujetos de la Educación",
            "Trabajo de Campo II"
        ],
        "3° Año": [
            "Macroeconomía II (Dinero, Crédito y Bancos)",
            "Economía Internacional y Comercio Exterior",
            "Finanzas Públicas y Política Fiscal",
            "Didáctica de la Economía I",
            "Historia Social de la Educación",
            "Derechos Humanos, Sociedad y Estado",
            "Construcción de la Práctica Docente I"
        ],
        "4° Año": [
            "Desarrollo Económico y Estructura Productiva Argentina",
            "Economía Política y Distribución del Ingreso",
            "Formulación y Evaluación de Proyectos",
            "Didáctica de la Economía II",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente",
            "Residencia y Construcción de la Práctica Docente II"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Economía Social y Cooperativismo",
            "Seminario de Integración Económica Regional",
            "Taller de Investigación Económica"
        ]
    },
    "Profesorado de Ciencias de la Administración": {
        "1° Año": [
            "Teoría de la Administración I (Fundamentos y Escuelas)",
            "Principios de Economía y Mercado",
            "Matemática Aplicada a la Administración",
            "Sistemas de Información y Contabilidad Básica",
            "Pedagogía",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Trabajo de Campo I"
        ],
        "2° Año": [
            "Teoría de la Administración II (Procesos Organizacionales)",
            "Contabilidad Financiera y de Costos",
            "Estadística Aplicada a las Organizaciones",
            "Derecho Aplicado a las Organizaciones",
            "Didáctica General",
            "Psicología Educacional",
            "Sujetos de la Educación",
            "Trabajo de Campo II"
        ],
        "3° Año": [
            "Administración de Recursos Humanos y Relaciones Laborales",
            "Administración Financiera y Presupuestaria",
            "Comercialización y Marketing Estratégico",
            "Didáctica de la Administración I",
            "Historia Social de la Educación",
            "Derechos Humanos, Sociedad y Estado",
            "Construcción de la Práctica Docente I"
        ],
        "4° Año": [
            "Planeamiento Estratégico y Dirección General",
            "Gestión de Organizaciones Públicas, Privadas y del Tercer Sector",
            "Emprendedorismo y Formulación de Proyectos",
            "Didáctica de la Administración II",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente",
            "Residencia y Construcción de la Práctica Docente II"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Innovación y Gestión Tecnológica",
            "Seminario de Responsabilidad Social Organizacional",
            "Taller de Simulación de Negocios y Práctica Profesional"
        ]
    },
    "Profesorado de Francés": {
        "1° Año": [
            "Lengua Francesa I",
            "Fonética y Fonología Francesa I",
            "Gramática Francesa I",
            "Práctica de Laboratorio y Expresión Oral I",
            "Civilización y Cultura Francesa I",
            "Pedagogía",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Construcción de la Práctica Docente en Nivel Inicial y Primario I"
        ],
        "2° Año": [
            "Lengua Francesa II",
            "Fonética y Fonología Francesa II",
            "Gramática Francesa II",
            "Literatura Francesa I (Siglos XVI al XVIII)",
            "Civilización y Cultura Francesa II",
            "Didáctica General",
            "Psicología Educacional",
            "Sujetos de la Educación",
            "Construcción de la Práctica Docente en Nivel Inicial y Primario II y Residencia"
        ],
        "3° Año": [
            "Lengua Francesa III",
            "Lingüística Aplicada al Francés",
            "Literatura Francesa II (Siglo XIX)",
            "Didáctica Especial y Construcción de la Práctica Docente en Nivel Medio I",
            "Historia Social de la Educación",
            "Derechos Humanos, Sociedad y Estado"
        ],
        "4° Año": [
            "Lengua Francesa IV",
            "Literatura Francesa y Francófona Contemporánea (Siglo XX y XXI)",
            "Estudios Francófonos y Diversidad Cultural",
            "Didáctica Especial y Residencia en Nivel Medio y Superior",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Literatura y Cine Francófono",
            "Seminario de Lingüística y Didáctica del FLE",
            "Taller de Traducción Francés-Español"
        ]
    },
    "Profesorado de Italiano": {
        "1° Año": [
            "Lengua Italiana I",
            "Fonética y Fonología Italiana I",
            "Gramática Italiana I",
            "Práctica de Laboratorio y Dicción I",
            "Cultura y Civilización Italiana I",
            "Pedagogía",
            "Lectura, Escritura y Oralidad I (LEO 1)",
            "Construcción de la Práctica Docente en Nivel Inicial y Primario I"
        ],
        "2° Año": [
            "Lengua Italiana II",
            "Fonética y Fonología Italiana II",
            "Gramática Italiana II",
            "Literatura Italiana I (Orígenes, Dante, Petrarca, Boccaccio)",
            "Cultura y Civilización Italiana II",
            "Didáctica General",
            "Psicología Educacional",
            "Sujetos de la Educación",
            "Construcción de la Práctica Docente en Nivel Inicial y Primario II y Residencia"
        ],
        "3° Año": [
            "Lengua Italiana III",
            "Lingüística Aplicada al Italiano",
            "Literatura Italiana II (Del Renacimiento al Romanticismo)",
            "Didáctica Especial y Construcción de la Práctica Docente en Nivel Medio I",
            "Historia Social de la Educación",
            "Derechos Humanos, Sociedad y Estado"
        ],
        "4° Año": [
            "Lengua Italiana IV",
            "Literatura Italiana Contemporánea (Siglos XX y XXI)",
            "Estudios Culturales de Italia y la Emigración Italiana en Argentina",
            "Didáctica Especial y Residencia en Nivel Medio y Superior",
            "Educación Sexual Integral (ESI)",
            "Ética y Trabajo Docente"
        ],
        "Tramo Superior / Seminarios": [
            "Seminario de Literatura Italiana y Traducción",
            "Seminario de Didáctica del Italiano como Lengua Extranjera",
            "Taller de Cultura y Cine Italiano"
        ]
    }
}

CAREER_FILENAMES = {
    "Profesorado de Filosofía": "Mapa_de_Carrera_Filosofia.xlsx",
    "Profesorado de Historia": "Mapa_de_Carrera_Historia.xlsx",
    "Profesorado de Lengua y Literatura": "Mapa_de_Carrera_Lengua_y_Literatura.xlsx",
    "Profesorado de Inglés": "Mapa_de_Carrera_Ingles.xlsx",
    "Profesorado de Matemática": "Mapa_de_Carrera_Matematica.xlsx",
    "Profesorado de Biología": "Mapa_de_Carrera_Biologia.xlsx",
    "Profesorado de Física": "Mapa_de_Carrera_Fisica.xlsx",
    "Profesorado de Química": "Mapa_de_Carrera_Quimica.xlsx",
    "Profesorado de Geografía": "Mapa_de_Carrera_Geografia.xlsx",
    "Profesorado de Informática": "Mapa_de_Carrera_Informatica.xlsx",
    "Profesorado de Psicología": "Mapa_de_Carrera_Psicologia.xlsx",
    "Profesorado de Ciencias de la Educación": "Mapa_de_Carrera_Ciencias_de_la_Educacion.xlsx",
    "Profesorado de Ciencias Jurídicas": "Mapa_de_Carrera_Ciencias_Juridicas.xlsx",
    "Profesorado de Ciencia Política": "Mapa_de_Carrera_Ciencia_Politica.xlsx",
    "Profesorado de Economía": "Mapa_de_Carrera_Economia.xlsx",
    "Profesorado de Ciencias de la Administración": "Mapa_de_Carrera_Ciencias_de_la_Administracion.xlsx",
    "Profesorado de Francés": "Mapa_de_Carrera_Frances.xlsx",
    "Profesorado de Italiano": "Mapa_de_Carrera_Italiano.xlsx"
}

def generate_all_excel_files(base_dir=r"d:\IA\AltilloJVG"):
    output_dir = os.path.join(base_dir, "data", "mapas")
    os.makedirs(output_dir, exist_ok=True)

    # Eliminar plantilla genérica
    generic_path = os.path.join(output_dir, "Mapa_de_Carrera_General_Plantilla.xlsx")
    if os.path.exists(generic_path):
        try:
            os.remove(generic_path)
            print("Eliminada plantilla generica previa.")
        except Exception:
            pass

    for career_name, years_data in MASTER_CURRICULUM.items():
        filename = CAREER_FILENAMES[career_name]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mapa de Carrera"

        # Colores
        navy_fill = PatternFill(start_color="0B2545", end_color="0B2545", fill_type="solid")
        cyan_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
        gold_fill = PatternFill(start_color="D99B26", end_color="D99B26", fill_type="solid")
        light_blue_fill = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
        light_gray_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        section_fill = PatternFill(start_color="134074", end_color="134074", fill_type="solid")
        green_soft = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

        font_title = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        font_subtitle = Font(name="Arial", size=10, italic=True, color="BAE6FD")
        font_section = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        font_regular = Font(name="Arial", size=10)
        font_bold = Font(name="Arial", size=10, bold=True)
        font_indicator = Font(name="Arial", size=11, bold=True, color="0B2545")

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # 1. Encabezado
        ws.merge_cells("A1:G1")
        ws["A1"] = f"MAPA DE CARRERA | {career_name.upper()}"
        ws["A1"].font = font_title
        ws["A1"].fill = navy_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:G2")
        ws["A2"] = "La Caravana + Estudiantes Independientes | Altillo JVG - ISP Joaquín V. González"
        ws["A2"].font = font_subtitle
        ws["A2"].fill = navy_fill
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        # 2. Encabezados de Columnas
        headers = [
            "Materia / Asignatura Oficial",
            "Cursada Regular",
            "Final Aprobado",
            "Año / Tramo",
            "Condición Cursada",
            "Condición Final",
            "Estado General"
        ]
        ws.row_dimensions[4].height = 26
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = cyan_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # Validación SI/NO
        dv = DataValidation(type="list", formula1='"SI,NO"', allow_blank=True)
        dv.error = "Por favor, elegí 'SI' o 'NO' de la lista desplegable."
        dv.errorTitle = "Opción inválida"
        ws.add_data_validation(dv)

        current_row = 5
        materia_rows = []

        for year_name, subjects in years_data.items():
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            sec_c = ws.cell(row=current_row, column=1, value=f"📌 {year_name.upper()}")
            sec_c.font = font_section
            sec_c.fill = section_fill
            sec_c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[current_row].height = 22
            current_row += 1

            for subj in subjects:
                row = current_row
                materia_rows.append(row)

                # A: Nombre Materia
                c_a = ws.cell(row=row, column=1, value=subj)
                c_a.font = font_regular
                c_a.border = thin_border
                c_a.alignment = Alignment(horizontal="left", vertical="center", indent=1)

                # B: Cursada Regular
                c_b = ws.cell(row=row, column=2, value="NO")
                c_b.font = font_bold
                c_b.alignment = Alignment(horizontal="center", vertical="center")
                c_b.border = thin_border
                c_b.fill = light_blue_fill
                dv.add(c_b)

                # C: Final Aprobado
                c_c = ws.cell(row=row, column=3, value="NO")
                c_c.font = font_bold
                c_c.alignment = Alignment(horizontal="center", vertical="center")
                c_c.border = thin_border
                c_c.fill = light_blue_fill
                dv.add(c_c)

                # D: Año
                c_d = ws.cell(row=row, column=4, value=year_name)
                c_d.font = font_regular
                c_d.alignment = Alignment(horizontal="center", vertical="center")
                c_d.border = thin_border

                # E: Condición Cursada
                c_e = ws.cell(row=row, column=5, value=f'=IF(B{row}="SI", "REGULAR", "PENDIENTE")')
                c_e.font = font_regular
                c_e.alignment = Alignment(horizontal="center", vertical="center")
                c_e.border = thin_border

                # F: Condición Final
                c_f = ws.cell(row=row, column=6, value=f'=IF(C{row}="SI", "APROBADO", IF(B{row}="SI", "ADEUDA FINAL", "ADEUDA"))')
                c_f.font = font_regular
                c_f.alignment = Alignment(horizontal="center", vertical="center")
                c_f.border = thin_border

                # G: Estado General
                c_g = ws.cell(row=row, column=7, value=f'=IF(C{row}="SI", "🟢 Aprobada", IF(B{row}="SI", "🟡 Regular", "🟣 Pendiente"))')
                c_g.font = font_bold
                c_g.alignment = Alignment(horizontal="center", vertical="center")
                c_g.border = thin_border

                ws.row_dimensions[row].height = 20
                current_row += 1

        total_materias = len(materia_rows)
        first_row = materia_rows[0]
        last_row = materia_rows[-1]

        # 3. Panel Lateral
        ws["I1"] = "CONTROL DE TRAYECTORIA"
        ws["I1"].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        ws["I1"].fill = navy_fill
        ws["I1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("I1:K1")

        stat_items = [
            ("Total de Materias del Plan", total_materias, False),
            ("Materias con Cursada Regular", f'=COUNTIF(B{first_row}:B{last_row}, "SI")', False),
            ("Finales / Materias Aprobadas", f'=COUNTIF(C{first_row}:C{last_row}, "SI")', False),
            ("Finales Adeudados", f'=COUNTIF(F{first_row}:F{last_row}, "ADEUDA FINAL")', False),
            ("% Cursadas Completadas", f'=COUNTIF(B{first_row}:B{last_row}, "SI")/{total_materias}', True),
            ("% Carrera Aprobada (Finales)", f'=COUNTIF(C{first_row}:C{last_row}, "SI")/{total_materias}', True),
        ]

        for s_idx, (label, formula_val, is_pct) in enumerate(stat_items, 2):
            ws.cell(row=s_idx, column=9, value=label).font = font_bold
            ws.cell(row=s_idx, column=9).fill = light_gray_fill
            ws.cell(row=s_idx, column=9).border = thin_border

            v_cell = ws.cell(row=s_idx, column=10, value=formula_val)
            v_cell.font = font_indicator
            v_cell.alignment = Alignment(horizontal="center", vertical="center")
            v_cell.border = thin_border
            v_cell.fill = green_soft if is_pct else light_blue_fill
            if is_pct:
                v_cell.number_format = '0.0%'

            ws.row_dimensions[s_idx].height = 22

        # 4. Instrucciones
        inst_row = 9
        ws.cell(row=inst_row, column=9, value="INSTRUCCIONES DE USO").font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        ws.cell(row=inst_row, column=9).fill = gold_fill
        ws.cell(row=inst_row, column=9).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=inst_row, start_column=9, end_row=inst_row, end_column=10)

        instrucciones = [
            "1. Hace clic en la celda y elegi 'SI' o 'NO' desde el menu desplegable.",
            "2. Columna B: Marca 'SI' si tenes la cursada regular aprobada.",
            "3. Columna C: Marca 'SI' si aprobaste el final o promocionaste.",
            "4. Las columnas E, F y G se actualizan automaticamente en tiempo real.",
            "5. El panel lateral calcula tu avance y porcentaje de carrera en vivo.",
            "6. Podes subir el archivo a tu Google Drive personal para editarlo online."
        ]

        for idx_i, inst in enumerate(instrucciones, inst_row + 1):
            cell = ws.cell(row=idx_i, column=9, value=inst)
            cell.font = Font(name="Arial", size=8.5, italic=True)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.merge_cells(start_row=idx_i, start_column=9, end_row=idx_i, end_column=10)
            ws.row_dimensions[idx_i].height = 20

        # Ancho Columnas
        ws.column_dimensions["A"].width = 54
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 24
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 4
        ws.column_dimensions["I"].width = 30
        ws.column_dimensions["J"].width = 14
        ws.column_dimensions["K"].width = 4

        out_path = os.path.join(output_dir, filename)
        wb.save(out_path)
        print(f"Generado mapa oficial curado: {filename} ({total_materias} materias)")

    # Actualizar upload_form.js con el curriculum curado
    upload_form_path = os.path.join(base_dir, "js", "upload_form.js")
    if os.path.exists(upload_form_path):
        with open(upload_form_path, "r", encoding="utf-8") as f:
            text = f.read()
        idx = text.find("const CURRICULUM_DATA = ")
        end_idx = text.find(";\n\ndocument.addEventListener")
        if idx != -1 and end_idx != -1:
            json_str = json.dumps(MASTER_CURRICULUM, ensure_ascii=False, indent=2)
            new_text = text[:idx + len("const CURRICULUM_DATA = ")] + json_str + text[end_idx:]
            with open(upload_form_path, "w", encoding="utf-8") as f:
                f.write(new_text)
            print("js/upload_form.js actualizado con curriculum curado.")

if __name__ == "__main__":
    generate_all_excel_files()
